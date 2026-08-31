"""Excel export service for Application Tracker using openpyxl."""

import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.application import ApplicationDocument


class ExportService:
    """Generates professional Excel workbooks for job applications."""

    @staticmethod
    async def export_applications_to_excel() -> io.BytesIO:
        """Export all applications into a styled XLSX spreadsheet."""
        applications = await ApplicationDocument.find_all().sort(-ApplicationDocument.created_at).to_list()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Applications"

        # Enable gridlines
        ws.views.sheetView[0].showGridLines = True

        # Define Headers
        headers = [
            "Company",
            "Role",
            "Status",
            "Location",
            "Salary Range",
            "Date Discovered",
            "Date Applied",
            "Interview Date",
            "AI Relevance Score",
            "Tags",
            "Resume Attached",
            "Matched Skills",
            "Missing Skills",
            "Notes",
            "Job Link",
        ]

        # Style definitions
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Calibri", size=10, color="000000")
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

        # Write Header Row
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[1].height = 28

        # Status color mappings
        status_colors = {
            "discovered": "E0E7FF",
            "applied": "DBEAFE",
            "responded": "CFFAFE",
            "interview_scheduled": "D1FAE5",
            "offer": "FEF3C7",
            "rejected": "FEE2E2",
            "ghosted": "F1F5F9",
        }

        # Write Application Rows
        for row_idx, app in enumerate(applications, start=2):
            date_disc = app.date_discovered.strftime("%Y-%m-%d") if app.date_discovered else ""
            date_app = app.date_applied.strftime("%Y-%m-%d") if app.date_applied else ""
            date_int = app.interview_date.strftime("%Y-%m-%d %H:%M") if app.interview_date else ""

            row_data = [
                app.company,
                app.role,
                app.status.value.replace("_", " ").title(),
                app.location or "",
                app.salary_range or "",
                date_disc,
                date_app,
                date_int,
                f"{app.relevance_score:.0f}%" if app.relevance_score is not None else "N/A",
                ", ".join(app.tags) if app.tags else "",
                app.resume_filename or "No",
                ", ".join(app.matched_skills) if app.matched_skills else "",
                ", ".join(app.missing_skills) if app.missing_skills else "",
                app.notes or "",
                app.job_url or "",
            ]

            ws.append(row_data)

            # Apply zebra styling and borders
            is_even = (row_idx % 2 == 0)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = regular_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

                if is_even and col_idx != 3:
                    cell.fill = zebra_fill

                # Special status column coloring
                if col_idx == 3:
                    bg_color = status_colors.get(app.status.value, "FFFFFF")
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[row_idx].height = 22

        # Auto-fit Column Widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

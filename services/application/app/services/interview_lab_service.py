"""Business logic and database service layer for Interview Lab."""

from datetime import datetime
import io
import random
from beanie import PydanticObjectId
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import structlog

from app.models.interview_lab import (
    ExperienceDocument,
    FlashCardDocument,
    QuestionDocument,
    TagDocument,
)
from shared.schemas.interview_lab import (
    ExperienceCreate,
    ExperienceResponse,
    ExperienceUpdate,
    FlashCardCreate,
    FlashCardResponse,
    FlashCardUpdate,
    QuestionCreate,
    QuestionResponse,
    QuestionUpdate,
    TagCreate,
    TagResponse,
    TagUpdate,
)

logger = structlog.get_logger()


class InterviewLabService:
    """Service handling CRUD operations and Excel reports for Interview Lab entities."""

    # -------------------------------------------------------------------------
    # Tags
    # -------------------------------------------------------------------------
    @staticmethod
    async def list_tags() -> list[TagResponse]:
        tags = await TagDocument.find_all().sort("-created_at").to_list()
        return [t.to_response_dto() for t in tags]

    @staticmethod
    async def create_tag(payload: TagCreate) -> TagResponse:
        clean_name = payload.name.strip()
        existing = await TagDocument.find_one(TagDocument.name == clean_name)
        if existing:
            return existing.to_response_dto()

        doc = TagDocument(name=clean_name, color=payload.color or "#c25e2e")
        await doc.insert()
        logger.info("interview_tag_created", name=clean_name)
        return doc.to_response_dto()

    @staticmethod
    async def update_tag(tag_id: str, payload: TagUpdate) -> TagResponse | None:
        try:
            oid = PydanticObjectId(tag_id)
        except Exception:
            return None

        doc = await TagDocument.get(oid)
        if not doc:
            return None

        if payload.name is not None:
            doc.name = payload.name.strip()
        if payload.color is not None:
            doc.color = payload.color

        await doc.save()
        return doc.to_response_dto()

    @staticmethod
    async def delete_tag(tag_id: str) -> bool:
        try:
            oid = PydanticObjectId(tag_id)
        except Exception:
            return False

        doc = await TagDocument.get(oid)
        if not doc:
            return False

        await doc.delete()
        logger.info("interview_tag_deleted", tag_id=tag_id)
        return True

    # -------------------------------------------------------------------------
    # Questions
    # -------------------------------------------------------------------------
    @staticmethod
    async def list_questions(
        application_id: str | None = None,
        company: str | None = None,
        difficulty: str | None = None,
        tag: str | None = None,
        search: str | None = None,
        limit: int = 200,
        offset: int = 0,
    ) -> list[QuestionResponse]:
        query = QuestionDocument.find()

        if application_id:
            query = query.find(QuestionDocument.application_id == application_id)
        elif company:
            query = query.find({"company": {"$regex": company, "$options": "i"}})

        if difficulty and difficulty.lower() != "all":
            query = query.find(QuestionDocument.difficulty == difficulty)

        if tag and tag.lower() != "all":
            query = query.find({"tags": tag})

        if search:
            regex_search = {"$regex": search, "$options": "i"}
            query = query.find(
                {
                    "$or": [
                        {"title": regex_search},
                        {"topic": regex_search},
                        {"description": regex_search},
                        {"notes": regex_search},
                    ]
                }
            )

        docs = await query.sort("-updated_at").skip(offset).limit(limit).to_list()
        return [q.to_response_dto() for q in docs]

    @staticmethod
    async def get_question(question_id: str) -> QuestionResponse | None:
        try:
            oid = PydanticObjectId(question_id)
        except Exception:
            return None

        doc = await QuestionDocument.get(oid)
        return doc.to_response_dto() if doc else None

    @staticmethod
    async def create_question(payload: QuestionCreate) -> QuestionResponse:
        doc = QuestionDocument(
            title=payload.title,
            description=payload.description,
            difficulty=payload.difficulty,
            topic=payload.topic,
            application_id=payload.application_id,
            company=payload.company,
            role=payload.role,
            solutions=payload.solutions,
            links=payload.links,
            tags=payload.tags,
            notes=payload.notes,
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        )
        await doc.insert()
        logger.info("interview_question_created", title=doc.title)
        return doc.to_response_dto()

    @staticmethod
    async def update_question(question_id: str, payload: QuestionUpdate) -> QuestionResponse | None:
        try:
            oid = PydanticObjectId(question_id)
        except Exception:
            return None

        doc = await QuestionDocument.get(oid)
        if not doc:
            return None

        update_dict = payload.model_dump(exclude_unset=True)
        for key, val in update_dict.items():
            setattr(doc, key, val)

        doc.updated_at = datetime.utcnow()
        await doc.save()
        logger.info("interview_question_updated", question_id=question_id)
        return doc.to_response_dto()

    @staticmethod
    async def delete_question(question_id: str) -> bool:
        try:
            oid = PydanticObjectId(question_id)
        except Exception:
            return False

        doc = await QuestionDocument.get(oid)
        if not doc:
            return False

        await doc.delete()
        logger.info("interview_question_deleted", question_id=question_id)
        return True

    # -------------------------------------------------------------------------
    # Experiences
    # -------------------------------------------------------------------------
    @staticmethod
    async def list_experiences(
        application_id: str | None = None,
        company: str | None = None,
        tag: str | None = None,
        min_rating: int | None = None,
        limit: int = 100,
        offset: int = 0,
    ) -> list[ExperienceResponse]:
        query = ExperienceDocument.find()

        if application_id:
            query = query.find(ExperienceDocument.application_id == application_id)
        elif company:
            query = query.find({"company": {"$regex": company, "$options": "i"}})

        if tag and tag.lower() != "all":
            query = query.find({"tags": tag})

        if min_rating:
            query = query.find(ExperienceDocument.rating >= min_rating)

        docs = await query.sort("-created_at").skip(offset).limit(limit).to_list()
        return [exp.to_response_dto() for exp in docs]

    @staticmethod
    async def get_experience(exp_id: str) -> ExperienceResponse | None:
        try:
            oid = PydanticObjectId(exp_id)
        except Exception:
            return None

        doc = await ExperienceDocument.get(oid)
        return doc.to_response_dto() if doc else None

    @staticmethod
    async def create_experience(payload: ExperienceCreate) -> ExperienceResponse:
        doc = ExperienceDocument(
            company=payload.company,
            role=payload.role,
            application_id=payload.application_id,
            date=payload.date or datetime.utcnow(),
            interview_process=payload.interview_process,
            questions_asked=payload.questions_asked,
            rating=payload.rating,
            overall_notes=payload.overall_notes,
            links=payload.links,
            tags=payload.tags,
            outcome=payload.outcome,
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        )
        await doc.insert()
        logger.info("interview_experience_created", company=doc.company, role=doc.role)
        return doc.to_response_dto()

    @staticmethod
    async def update_experience(exp_id: str, payload: ExperienceUpdate) -> ExperienceResponse | None:
        try:
            oid = PydanticObjectId(exp_id)
        except Exception:
            return None

        doc = await ExperienceDocument.get(oid)
        if not doc:
            return None

        update_dict = payload.model_dump(exclude_unset=True)
        for key, val in update_dict.items():
            setattr(doc, key, val)

        doc.updated_at = datetime.utcnow()
        await doc.save()
        logger.info("interview_experience_updated", exp_id=exp_id)
        return doc.to_response_dto()

    @staticmethod
    async def delete_experience(exp_id: str) -> bool:
        try:
            oid = PydanticObjectId(exp_id)
        except Exception:
            return False

        doc = await ExperienceDocument.get(oid)
        if not doc:
            return False

        await doc.delete()
        logger.info("interview_experience_deleted", exp_id=exp_id)
        return True

    # -------------------------------------------------------------------------
    # Flash Cards
    # -------------------------------------------------------------------------
    @staticmethod
    async def list_flashcards(
        application_id: str | None = None,
        company: str | None = None,
        tag: str | None = None,
        difficulty: str | None = None,
        limit: int = 200,
        offset: int = 0,
    ) -> list[FlashCardResponse]:
        query = FlashCardDocument.find()

        if application_id:
            query = query.find(FlashCardDocument.application_id == application_id)
        elif company:
            query = query.find({"company": {"$regex": company, "$options": "i"}})

        if tag and tag.lower() != "all":
            query = query.find({"tags": tag})

        if difficulty and difficulty.lower() != "all":
            query = query.find(FlashCardDocument.difficulty == difficulty)

        docs = await query.sort("-updated_at").skip(offset).limit(limit).to_list()
        return [fc.to_response_dto() for fc in docs]

    @staticmethod
    async def get_flashcard(card_id: str) -> FlashCardResponse | None:
        try:
            oid = PydanticObjectId(card_id)
        except Exception:
            return None

        doc = await FlashCardDocument.get(oid)
        return doc.to_response_dto() if doc else None

    @staticmethod
    async def get_study_deck(
        application_id: str | None = None,
        count: int = 20,
        tag: str | None = None,
        difficulty: str | None = None,
        shuffle: bool = True,
    ) -> list[FlashCardResponse]:
        """Fetch flashcards for active study session."""
        query = FlashCardDocument.find()

        if application_id:
            query = query.find(FlashCardDocument.application_id == application_id)

        if tag and tag.lower() != "all":
            query = query.find({"tags": tag})

        if difficulty and difficulty.lower() != "all":
            query = query.find(FlashCardDocument.difficulty == difficulty)

        docs = await query.to_list()
        if shuffle:
            random.shuffle(docs)

        selected = docs[:count] if count > 0 else docs
        return [fc.to_response_dto() for fc in selected]

    @staticmethod
    async def create_flashcard(payload: FlashCardCreate) -> FlashCardResponse:
        doc = FlashCardDocument(
            front=payload.front,
            back=payload.back,
            application_id=payload.application_id,
            company=payload.company,
            tags=payload.tags,
            links=payload.links,
            difficulty=payload.difficulty,
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        )
        await doc.insert()
        logger.info("flashcard_created", front_snippet=doc.front[:30])
        return doc.to_response_dto()

    @staticmethod
    async def update_flashcard(card_id: str, payload: FlashCardUpdate) -> FlashCardResponse | None:
        try:
            oid = PydanticObjectId(card_id)
        except Exception:
            return None

        doc = await FlashCardDocument.get(oid)
        if not doc:
            return None

        update_dict = payload.model_dump(exclude_unset=True)
        for key, val in update_dict.items():
            setattr(doc, key, val)

        doc.updated_at = datetime.utcnow()
        await doc.save()
        logger.info("flashcard_updated", card_id=card_id)
        return doc.to_response_dto()

    @staticmethod
    async def delete_flashcard(card_id: str) -> bool:
        try:
            oid = PydanticObjectId(card_id)
        except Exception:
            return False

        doc = await FlashCardDocument.get(oid)
        if not doc:
            return False

        await doc.delete()
        logger.info("flashcard_deleted", card_id=card_id)
        return True

    # -------------------------------------------------------------------------
    # Excel Export
    # -------------------------------------------------------------------------
    @staticmethod
    async def export_excel(application_id: str | None = None, company: str | None = None) -> bytes:
        """Generate a multi-sheet styled Excel (.xlsx) file containing Experiences, Questions, and Flashcards."""
        exps = await InterviewLabService.list_experiences(application_id=application_id, company=company, limit=500)
        questions = await InterviewLabService.list_questions(application_id=application_id, company=company, limit=500)
        flashcards = await InterviewLabService.list_flashcards(application_id=application_id, company=company, limit=500)

        wb = openpyxl.Workbook()

        # Styles
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Calibri", size=10)
        bold_font = Font(name="Calibri", size=10, bold=True)
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

        # ---------------------------------------------------------------------
        # Sheet 1: Interview Experiences
        # ---------------------------------------------------------------------
        ws_exp = wb.active
        ws_exp.title = "Interview Experiences"

        exp_headers = [
            "Company",
            "Role",
            "Date",
            "Outcome",
            "Self-Rating (1-10)",
            "Process Pipeline Rounds",
            "Questions Asked & Answers",
            "Retrospective & Notes",
            "Tags",
            "Links",
        ]
        ws_exp.append(exp_headers)
        for col_idx in range(1, len(exp_headers) + 1):
            cell = ws_exp.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx, exp in enumerate(exps, start=2):
            pipeline_str = " ➔ ".join([f"R{s.round_number}: {s.round_type} ({s.description})" if s.description else f"R{s.round_number}: {s.round_type}" for s in exp.interview_process])
            qa_str = "\n".join([f"Q: {qa.question}\nA: {qa.answer} [{qa.category}]" for qa in exp.questions_asked])
            links_str = ", ".join([f"{l.label}: {l.url}" for l in exp.links])
            date_str = exp.date.strftime("%Y-%m-%d") if exp.date else ""

            row_data = [
                exp.company,
                exp.role,
                date_str,
                exp.outcome or "Pending",
                f"{exp.rating}/10",
                pipeline_str,
                qa_str,
                exp.overall_notes,
                ", ".join(exp.tags),
                links_str,
            ]
            ws_exp.append(row_data)

            for col_idx in range(1, len(row_data) + 1):
                cell = ws_exp.cell(row=row_idx, column=col_idx)
                cell.font = regular_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if row_idx % 2 == 1:
                    cell.fill = zebra_fill

        # ---------------------------------------------------------------------
        # Sheet 2: DSA & LeetCode Questions
        # ---------------------------------------------------------------------
        ws_q = wb.create_sheet(title="DSA & LeetCode Questions")
        q_headers = [
            "Title",
            "Difficulty",
            "Topic",
            "Problem Summary",
            "Solutions Count",
            "Solution Approaches & Complexities",
            "Key Notes",
            "Tags",
            "Links",
        ]
        ws_q.append(q_headers)
        for col_idx in range(1, len(q_headers) + 1):
            cell = ws_q.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx, q in enumerate(questions, start=2):
            solutions_str = "\n\n".join([
                f"[{sol.label}] ({sol.language})\nTime: {sol.time_complexity} | Space: {sol.space_complexity}\nIntuition: {sol.explanation}\nCode:\n{sol.code}"
                for sol in q.solutions
            ])
            links_str = ", ".join([f"{l.label}: {l.url}" for l in q.links])

            row_data = [
                q.title,
                q.difficulty,
                q.topic,
                q.description,
                len(q.solutions),
                solutions_str,
                q.notes,
                ", ".join(q.tags),
                links_str,
            ]
            ws_q.append(row_data)

            for col_idx in range(1, len(row_data) + 1):
                cell = ws_q.cell(row=row_idx, column=col_idx)
                cell.font = regular_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if row_idx % 2 == 1:
                    cell.fill = zebra_fill

        # ---------------------------------------------------------------------
        # Sheet 3: Flashcards
        # ---------------------------------------------------------------------
        ws_fc = wb.create_sheet(title="Revision Flashcards")
        fc_headers = [
            "Front (Question / Prompt)",
            "Back (Answer / Explanation)",
            "Difficulty",
            "Tags",
            "Links",
        ]
        ws_fc.append(fc_headers)
        for col_idx in range(1, len(fc_headers) + 1):
            cell = ws_fc.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx, fc in enumerate(flashcards, start=2):
            links_str = ", ".join([f"{l.label}: {l.url}" for l in fc.links])
            row_data = [
                fc.front,
                fc.back,
                fc.difficulty,
                ", ".join(fc.tags),
                links_str,
            ]
            ws_fc.append(row_data)

            for col_idx in range(1, len(row_data) + 1):
                cell = ws_fc.cell(row=row_idx, column=col_idx)
                cell.font = regular_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if row_idx % 2 == 1:
                    cell.fill = zebra_fill

        # Auto-adjust column widths for all sheets
        for sheet in [ws_exp, ws_q, ws_fc]:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or "")
                    first_line = val_str.split("\n")[0]
                    max_len = max(max_len, len(first_line))
                sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream.getvalue()
